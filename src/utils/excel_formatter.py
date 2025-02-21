# src/utils/excel_formatter.py
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

class ExcelStyler:
   """Utility for applying consistent Excel styling with bilingual support"""
   
   @staticmethod
   def apply_header_style(cell, is_bilingual=True):
       """Apply header styling with optional bilingual formatting"""
       cell.font = Font(bold=True, size=12)
       cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
       cell.border = Border(
           bottom=Side(style='medium'),
           top=Side(style='medium'),
           left=Side(style='thin'),
           right=Side(style='thin')
       )
       cell.alignment = Alignment(
           horizontal='center', 
           vertical='center', 
           wrap_text=True
       )
       
       # Optional bilingual text processing
       if is_bilingual and isinstance(cell.value, str) and '/' in cell.value:
           cell.value = cell.value.strip()

   @staticmethod
   def apply_data_style(cell, is_metric=False, is_bilingual=True):
       """Apply data cell styling with optional metric and bilingual formatting"""
       cell.border = Border(
           bottom=Side(style='thin'),
           top=Side(style='thin'),
           left=Side(style='thin'),
           right=Side(style='thin')
       )
       if is_metric:
           cell.number_format = '0.000'
       cell.alignment = Alignment(
           horizontal='center', 
           vertical='center', 
           wrap_text=True
       )
       
       # Optional bilingual text processing for text values
       if is_bilingual and isinstance(cell.value, str) and '/' in cell.value:
           # Split and take first part (English)
           cell.value = cell.value.split('/')[0].strip()

   @staticmethod
   def adjust_column_width(worksheet, max_width=40):
       """Dynamically adjust column widths"""
       for column in worksheet.columns:
           max_length = 0
           column_letter = get_column_letter(column[0].column)
           for cell in column:
               try:
                   content_length = len(str(cell.value)) if cell.value is not None else 0
                   max_length = max(max_length, content_length)
               except:
                   pass
           adjusted_width = (max_length + 2) * 1.2
           worksheet.column_dimensions[column_letter].width = min(adjusted_width, max_width)

   @staticmethod
   def create_bilingual_label(english, arabic):
       """Create a bilingual label"""
       return f"{english} / {arabic}"