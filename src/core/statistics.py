# src/core/statistics.py
from openpyxl import Workbook
from openpyxl.styles import Font
from .alpha_analysis import CronbachAlphaAnalyzer
from ..utils.excel_formatter import ExcelStyler
import logging
import pandas as pd
import numpy as np

def calculate_split_half_reliability(data: pd.DataFrame, questions: list, logger=None):
   """
   Calculate split-half reliability
   
   :param data: DataFrame containing the data
   :param questions: List of question column names
   :return: Dictionary with split-half reliability metrics
   """
   if logger is None:
       logger = logging.getLogger()
       
   df = data[questions].apply(pd.to_numeric, errors='coerce')
   
   # Split questions into odd and even
   odd_questions = questions[::2]
   even_questions = questions[1::2]
   
   logger.info("\nSplit-Half Analysis:")
   logger.info(f"Odd questions: {odd_questions}")
   logger.info(f"Even questions: {even_questions}")
   
   # Calculate sums
   odd_sums = df[odd_questions].sum(axis=1)
   even_sums = df[even_questions].sum(axis=1)
   
   # Calculate correlations
   pearson_corr = odd_sums.corr(even_sums)
   spearman_brown = (2 * pearson_corr) / (1 + pearson_corr)
   
   logger.info(f"\nPearson correlation: {pearson_corr}")
   logger.info(f"Spearman-Brown coefficient: {spearman_brown}")
   
   return {
       'pearson_correlation': pearson_corr,
       'spearman_brown': spearman_brown,
       'odd_questions': odd_questions,
       'even_questions': even_questions
   }

def calculate_construct_validity(data: pd.DataFrame, questions: list, dimensions: dict, logger=None):
   """
   Calculate construct validity for dimensions
   
   :param data: DataFrame containing the data
   :param questions: List of question column names
   :param dimensions: Dictionary of dimensions with their questions
   :return: Dictionary with construct validity metrics
   """
   if logger is None:
       logger = logging.getLogger()
       
   df = data[questions].apply(pd.to_numeric, errors='coerce')
   
   # Calculate total scores
   total_scores = df.sum(axis=1)
   total_ranks = total_scores.rank(method='average')
   
   logger.info("\nConstruct Validity Analysis:")
   logger.info(f"Total questions: {len(questions)}")
   
   # Calculate dimension statistics
   dimension_stats = {}
   for dim_name, dim_questions in dimensions.items():
       # Ensure dim_questions are in the correct order and exist in the data
       valid_dim_questions = [q for q in dim_questions if q in data.columns]
       
       dim_scores = data[valid_dim_questions].apply(pd.to_numeric, errors='coerce').sum(axis=1)
       dim_ranks = dim_scores.rank(method='average')
       correlation = total_ranks.corr(dim_ranks, method='spearman')
       
       dimension_stats[dim_name] = {
           'scores': dim_scores,
           'ranks': dim_ranks,
           'correlation': correlation
       }
       
       logger.info(f"Dimension {dim_name} correlation with total: {correlation}")
   
   return {
       'total_scores': total_scores,
       'total_ranks': total_ranks,
       'dimensions': dimension_stats
   }

def export_statistics(filepath: str, data: pd.DataFrame, questions: list, dimensions: dict):
   """Export statistical analysis with detailed breakdown"""
   wb = Workbook()
   logger = logging.getLogger("ExcelAutoRanker")
   
   # Calculate overall alpha with deletion impacts
   total_alpha = CronbachAlphaAnalyzer.calculate_with_deletion_impact(data, questions, logger)
   
   # Calculate dimension alphas
   dimension_alphas = {}
   for dim_name, dim_questions in dimensions.items():
       dimension_alphas[dim_name] = CronbachAlphaAnalyzer.calculate(data, dim_questions, logger)
   
   # Split-half reliability
   split_half_stats = calculate_split_half_reliability(data, questions, logger)
   
   # Construct validity
   construct_stats = calculate_construct_validity(data, questions, dimensions, logger)
   
   # Total Cronbach's Alpha Sheet
   total_sheet = wb.active
   total_sheet.title = "Total Cronbach's Alpha"
   
   headers = [
       "Metric / المقياس",
       "Value / القيمة",
       "Interpretation / التفسير"
   ]
   
   for col, header in enumerate(headers, 1):
       cell = total_sheet.cell(row=1, column=col, value=header)
       ExcelStyler.apply_header_style(cell)
   
   # Add total alpha details
   row_data = [
       ("Total Items / العناصر الكلية", total_alpha['n_items'], ""),
       ("Cronbach's Alpha / معامل ألفا كرونباخ", total_alpha['alpha'], 
        "Excellent" if total_alpha['alpha'] > 0.9 else 
        "Good" if total_alpha['alpha'] > 0.8 else 
        "Acceptable" if total_alpha['alpha'] > 0.7 else "Poor")
   ]
   
   for row, (metric, value, interpretation) in enumerate(row_data, 2):
       total_sheet.cell(row=row, column=1, value=metric)
       total_sheet.cell(row=row, column=2, value=value)
       total_sheet.cell(row=row, column=3, value=interpretation)
   
   # Dimension Alphas Sheet
   dim_sheet = wb.create_sheet("Dimension Alphas")
   
   dim_headers = [
       "Dimension / البعد",
       "Alpha / معامل ألفا",
       "Items / العناصر",
       "Interpretation / التفسير"
   ]
   
   for col, header in enumerate(dim_headers, 1):
       cell = dim_sheet.cell(row=1, column=col, value=header)
       ExcelStyler.apply_header_style(cell)
   
   # Add dimension alpha details
   for row, (dim_name, dim_alpha) in enumerate(dimension_alphas.items(), 2):
       dim_sheet.cell(row=row, column=1, value=f"Dimension {dim_name}")
       dim_sheet.cell(row=row, column=2, value=dim_alpha['alpha'])
       dim_sheet.cell(row=row, column=3, value=dim_alpha['n_items'])
       dim_sheet.cell(row=row, column=4, value=
           "Excellent" if dim_alpha['alpha'] > 0.9 else 
           "Good" if dim_alpha['alpha'] > 0.8 else 
           "Acceptable" if dim_alpha['alpha'] > 0.7 else "Poor"
       )
   
   # Question Deletion Impact Sheet
   deletion_sheet = wb.create_sheet("Question Deletion Impact")
   
   deletion_headers = [
       "Question / السؤال",
       "Alpha if Deleted / معامل ألفا عند الحذف",
       "Alpha Change / التغير في معامل ألفا"
   ]
   
   for col, header in enumerate(deletion_headers, 1):
       cell = deletion_sheet.cell(row=1, column=col, value=header)
       ExcelStyler.apply_header_style(cell)
   
   # Add question deletion impacts
   for row, (question, impact) in enumerate(total_alpha['deletion_impacts'].items(), 2):
       deletion_sheet.cell(row=row, column=1, value=question)
       deletion_sheet.cell(row=row, column=2, value=impact['alpha_if_deleted'])
       deletion_sheet.cell(row=row, column=3, value=impact['alpha_change'])
   
   # Adjust column widths
   for sheet in wb.worksheets:
       ExcelStyler.adjust_column_width(sheet)
   
   # Save workbook
   wb.save(filepath)
   
   return {
       'total_alpha': total_alpha,
       'dimension_alphas': dimension_alphas,
       'split_half': split_half_stats,
       'construct_validity': construct_stats
   }