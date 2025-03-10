# src/core/statistics_manager.py
from openpyxl import Workbook
import pandas as pd
import logging
from typing import List, Dict
import os
from .statistics.cronbach_alpha import CronbachAlphaCalculator
from .statistics.split_half import SplitHalfCalculator
from .statistics.construct_validity import ConstructValidityCalculator
from .formatters.cronbach_formatter import CronbachFormatter
from .formatters.split_half_formatter import SplitHalfFormatter
from .formatters.construct_formatter import ConstructValidityFormatter

class StatisticsManager:
    def __init__(self, data: pd.DataFrame, questions: List[str], dimensions: Dict[str, List[str]]):
        self.logger = logging.getLogger('ExcelAutoRanker')
        self.data = data
        self.questions = questions
        self.dimensions = dimensions
        self.cronbach = CronbachAlphaCalculator()
        self.split_half = SplitHalfCalculator()
        self.construct_validity = ConstructValidityCalculator()
        
        self.logger.info("\n" + "="*80)
        self.logger.info("STATISTICS CALCULATION SETUP")
        self.logger.info(f"Total questions: {len(questions)}")
        self.logger.info(f"Total dimensions: {len(dimensions)}")
        self.logger.info("-"*80)
        self.logger.info("DIMENSIONS BREAKDOWN:")
        for dim_num, dim_cols in dimensions.items():
            self.logger.info(f"Dimension {dim_num} ({len(dim_cols)} questions):")
            self.logger.info(f"Questions: {','.join(dim_cols)}")
            self.logger.info("-"*40)

    def calculate_per_question_alpha(self) -> Dict:
        """Calculate Cronbach's Alpha excluding each question one at a time"""
        per_question_results = {}
        
        self.logger.info("Calculating per-question Cronbach's Alpha")
        
        # Calculate baseline alpha with all questions
        baseline_alpha = self.cronbach.calculate(self.data, self.questions)['alpha']
        
        for question in self.questions:
            # Create list of questions excluding current one
            remaining_questions = [q for q in self.questions if q != question]
            
            # Calculate alpha without this question
            alpha_without = self.cronbach.calculate(self.data, remaining_questions)['alpha']
            
            # Store results
            per_question_results[question] = {
                'alpha_if_deleted': alpha_without,
                'alpha_change': alpha_without - baseline_alpha
            }
            
            self.logger.debug(f"Question {question}: Alpha if deleted = {alpha_without:.4f}, Change = {alpha_without - baseline_alpha:.4f}")
        
        return per_question_results

    def calculate_per_question_construct_validity(self) -> Dict:
        """Calculate Construct Validity for each question within its dimension"""
        per_question_results = {}
        
        self.logger.info("Calculating per-question Construct Validity")
        
        try:
            # Convert to numeric and handle missing values
            df = self.data[self.questions].apply(pd.to_numeric, errors='coerce')
            df_clean = df.dropna()
            
            # Process each dimension
            for dim_num, dim_questions in self.dimensions.items():
                self.logger.debug(f"Processing dimension {dim_num}")
                
                # Calculate dimension total score
                dim_total = df_clean[dim_questions].sum(axis=1)
                
                # Calculate correlation for each question in the dimension
                for question in dim_questions:
                    # Calculate dimension total without current question
                    remaining_questions = [q for q in dim_questions if q != question]
                    dim_total_without = df_clean[remaining_questions].sum(axis=1)
                    
                    # Calculate correlation between question and dimension total (without the question)
                    question_scores = df_clean[question]
                    correlation = question_scores.corr(dim_total_without, method='spearman')
                    
                    # Store results
                    per_question_results[question] = {
                        'dimension': dim_num,
                        'correlation': correlation,
                        'interpretation': self._get_correlation_interpretation(correlation)
                    }
                    
                    self.logger.debug(f"Question {question}: Correlation = {correlation:.4f}")
            
            return per_question_results
            
        except Exception as e:
            self.logger.error(f"Error in per-question construct validity calculation: {str(e)}")
            self.logger.error("Full error details:", exc_info=True)
            return {}

    def _get_correlation_interpretation(self, correlation: float) -> str:
        """Get bilingual interpretation of correlation coefficient"""
        abs_corr = abs(correlation)
        if abs_corr > 0.7:
            return "Strong / قوي"
        elif abs_corr > 0.5:
            return "Moderate / متوسط"
        elif abs_corr > 0.3:
            return "Weak / ضعيف"
        else:
            return "Very Weak / ضعيف جداً"

    def format_per_question_results(self, ws):
        """Format per-question Cronbach's Alpha analysis results in worksheet"""
        # Headers
        headers = [
            "Question / السؤال",
            "Alpha if Deleted / معامل ألفا عند الحذف",
            "Change in Alpha / التغير في معامل ألفا",
            "Impact / التأثير"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            CronbachFormatter._apply_header_style(cell)
        
        # Calculate per-question results
        results = self.calculate_per_question_alpha()
        
        # Add data rows
        for row, (question, data) in enumerate(results.items(), 2):
            ws.cell(row=row, column=1, value=question)
            ws.cell(row=row, column=2, value=round(data['alpha_if_deleted'], 6))
            change = data['alpha_change']
            ws.cell(row=row, column=3, value=round(change, 6))
            
            # Add impact interpretation
            impact = ("Improves reliability / يحسن الثبات" if change > 0 else 
                     "Reduces reliability / يقلل الثبات" if change < 0 else 
                     "No impact / لا تأثير")
            ws.cell(row=row, column=4, value=impact)
            
            # Apply styling to all cells in row
            for col in range(1, 5):
                CronbachFormatter._apply_data_style(ws.cell(row=row, column=col))
        
        # Adjust column widths
        CronbachFormatter._adjust_column_widths(ws)

    def format_per_question_construct_validity(self, ws):
        """Format per-question construct validity results in worksheet"""
        # Headers
        headers = [
            "Question / السؤال",
            "Dimension / البعد",
            "Correlation / معامل الارتباط",
            "Interpretation / التفسير"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            ConstructValidityFormatter._apply_header_style(cell)
        
        # Calculate per-question results
        results = self.calculate_per_question_construct_validity()
        
        # Add data rows
        for row, (question, data) in enumerate(results.items(), 2):
            ws.cell(row=row, column=1, value=question)
            ws.cell(row=row, column=2, value=f"Dimension {data['dimension']}")
            ws.cell(row=row, column=3, value=round(data['correlation'], 6))
            ws.cell(row=row, column=4, value=data['interpretation'])
            
            # Apply styling to all cells in row
            for col in range(1, 5):
                ConstructValidityFormatter._apply_data_style(ws.cell(row=row, column=col))
        
        # Adjust column widths
        ConstructValidityFormatter._adjust_column_widths(ws)

    def analyze_and_export(self, output_dir: str = '/app/data/output') -> str:
        """Run statistical analysis and export to Excel"""
        try:
            # Ensure output directory exists
            os.makedirs(output_dir, exist_ok=True)
            
            # Create workbook
            wb = Workbook()
            
            # Calculate and format overall Cronbach's Alpha
            self.logger.info("Calculating total Cronbach's Alpha")
            total_alpha_results = self.cronbach.calculate(self.data, self.questions)
            self.logger.info(f"Total Cronbach's Alpha: {total_alpha_results.get('alpha', 'N/A')}")
            CronbachFormatter.format_results(wb, total_alpha_results)
            
            # Calculate and format Split-Half reliability
            self.logger.info("Calculating Split-Half reliability")
            split_half_results = self.split_half.calculate(self.data, self.questions)
            split_half_sheet = wb.create_sheet(title="Split Half")
            SplitHalfFormatter.format_results_to_sheet(split_half_sheet, split_half_results)
            
            # Calculate and format Construct Validity
            self.logger.info("Calculating Construct Validity")
            construct_validity_results = self.construct_validity.calculate(
                self.data, self.questions, self.dimensions
            )
            construct_validity_sheet = wb.create_sheet(title="Construct Validity")
            ConstructValidityFormatter.format_results_to_sheet(construct_validity_sheet, construct_validity_results)
            
            # Add per-question construct validity analysis sheet
            per_question_cv_sheet = wb.create_sheet(title="Question Construct Validity")
            self.format_per_question_construct_validity(per_question_cv_sheet)
            
            # Calculate and format dimensional statistics
            self.logger.info("Calculating dimensional statistics")
            for dim_num, dim_questions in self.dimensions.items():
                self.logger.debug(f"Processing dimension {dim_num} with {len(dim_questions)} questions")
                
                # Calculate both Cronbach's Alpha and Split-Half for each dimension
                dim_alpha_results = self.cronbach.calculate(self.data, dim_questions)
                dim_split_half_results = self.split_half.calculate(self.data, dim_questions)
                
                self.logger.info(f"Dimension {dim_num} Cronbach's Alpha: {dim_alpha_results.get('alpha', 'N/A')}")
                self.logger.info(f"Dimension {dim_num} Split-Half: {dim_split_half_results.get('spearman_brown', 'N/A')}")
                
                if dim_alpha_results['status'] == 'success':
                    # Create sheets for both analyses
                    alpha_sheet = wb.create_sheet(title=f"Dimension {dim_num} Alpha")
                    split_half_sheet = wb.create_sheet(title=f"Dimension {dim_num} Split")
                    
                    CronbachFormatter.format_results_to_sheet(alpha_sheet, dim_alpha_results)
                    SplitHalfFormatter.format_results_to_sheet(split_half_sheet, dim_split_half_results)
                else:
                    self.logger.error(f"Failed to calculate statistics for dimension {dim_num}")

            # Add per-question Cronbach's alpha analysis sheet
            per_question_sheet = wb.create_sheet(title="Question Analysis")
            self.format_per_question_results(per_question_sheet)
            
            # Save workbook
            output_file = os.path.join(output_dir, 'statistical_analysis.xlsx')
            wb.save(output_file)
            self.logger.info(f"Analysis exported to {output_file}")
            
            return output_file
            
        except Exception as e:
            self.logger.error(f"Error in analyze_and_export: {str(e)}")
            raise