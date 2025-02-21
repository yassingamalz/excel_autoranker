# src/core/formatters/cronbach_formatter.py
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from typing import Dict

class CronbachFormatter:
    """Formats Cronbach's Alpha results into Excel worksheet"""
    
    @staticmethod
    def format_results(wb: Workbook, results: Dict, prefix: str = "") -> None:
        """Format results to the active sheet with optional prefix"""
        ws = wb.active
        title_prefix = f"{prefix} " if prefix else ""
        ws.title = f"{title_prefix}Cronbach Alpha - معامل ألفا"
        CronbachFormatter.format_results_to_sheet(ws, results)

    @staticmethod
    def format_results_to_sheet(ws: Worksheet, results: Dict) -> None:
        """Format results to a specific worksheet"""
        # Apply headers
        headers = [
            "Metric / المقياس",
            "Value / القيمة",
            "Interpretation / التفسير"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            CronbachFormatter._apply_header_style(cell)
        
        # Add results
        row_data = [
            ("Total Items / إجمالي العناصر", 
             results['n_items'], 
             ""),
            ("Sum of Item Variances / مجموع تباينات الأسئلة",
             round(sum(results['item_variances'].values()), 6),
             ""),
            ("Total Score Variance / تباين المجموع الكلي",
             round(results['total_variance'], 6),
             ""),
            ("Cronbach's Alpha / معامل ألفا كرونباخ",
             round(results['alpha'], 6) if results['alpha'] is not None else "N/A",
             results.get('interpretation', ''))
        ]
        
        for row, (metric, value, interp) in enumerate(row_data, 2):
            ws.cell(row=row, column=1, value=metric)
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=3, value=interp)
            CronbachFormatter._apply_data_style(ws.cell(row=row, column=1))
            CronbachFormatter._apply_data_style(ws.cell(row=row, column=2))
            CronbachFormatter._apply_data_style(ws.cell(row=row, column=3))
            
        # Adjust column widths
        CronbachFormatter._adjust_column_widths(ws)
        
    @staticmethod
    def _apply_header_style(cell):
        """Apply consistent header styling"""
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        cell.border = Border(
            bottom=Side(style='medium'),
            top=Side(style='medium'),
            left=Side(style='thin'),
            right=Side(style='thin')
        )
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    @staticmethod
    def _apply_data_style(cell):
        """Apply consistent data cell styling"""
        cell.border = Border(
            bottom=Side(style='thin'),
            top=Side(style='thin'),
            left=Side(style='thin'),
            right=Side(style='thin')
        )
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    @staticmethod
    def _adjust_column_widths(ws):
        """Adjust column widths based on content"""
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = min(adjusted_width, 40)