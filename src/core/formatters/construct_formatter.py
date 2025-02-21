# src/core/formatters/construct_formatter.py
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from typing import Dict

class ConstructValidityFormatter:
    """Formats Construct Validity results into Excel worksheet"""
    
    @staticmethod
    def format_results(wb: Workbook, results: Dict, prefix: str = "") -> None:
        """Format results to the active sheet with optional prefix"""
        ws = wb.active
        title_prefix = f"{prefix} " if prefix else ""
        ws.title = f"{title_prefix}Construct Validity - الصدق البنائي"
        ConstructValidityFormatter.format_results_to_sheet(ws, results)

    @staticmethod
    def format_results_to_sheet(ws: Worksheet, results: Dict) -> None:
        """Format results to a specific worksheet"""
        # Apply headers for correlation summary
        headers = [
            "Dimension / البعد",
            "Correlation / معامل الارتباط",
            "Interpretation / التفسير"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            ConstructValidityFormatter._apply_header_style(cell)
        
        # Add correlation results
        correlations = results['correlations']
        interpretations = results['interpretation']
        
        for row, (dim_num, correlation) in enumerate(correlations.items(), 2):
            ws.cell(row=row, column=1, value=f"Dimension {dim_num} / البعد {dim_num}")
            ws.cell(row=row, column=2, value=round(correlation, 6))
            ws.cell(row=row, column=3, value=interpretations[dim_num])
            
            for col in range(1, 4):
                ConstructValidityFormatter._apply_data_style(ws.cell(row=row, column=col))
        
        # Add detailed scores table starting at row 8
        ws.cell(row=7, column=1, value="Participant Scores and Ranks / درجات وترتيب المشاركين")
        ws.cell(row=8, column=1, value="Participant / المشارك")
        ws.cell(row=8, column=2, value="Total Score / المجموع الكلي")
        ws.cell(row=8, column=3, value="Total Rank / الترتيب الكلي")
        
        col_offset = 4
        for dim_num in correlations.keys():
            ws.cell(row=8, column=col_offset, value=f"Dim {dim_num} Score / درجة البعد {dim_num}")
            ws.cell(row=8, column=col_offset + 1, value=f"Dim {dim_num} Rank / ترتيب البعد {dim_num}")
            col_offset += 2
        
        for col in range(1, col_offset):
            ConstructValidityFormatter._apply_header_style(ws.cell(row=8, column=col))
        
        # Add participant data
        total_scores = results['total_scores']
        total_ranks = results['total_ranks']
        dimension_results = results['dimension_results']
        
        for idx, (participant, total_score) in enumerate(total_scores.items(), 9):
            # Add participant and total scores/ranks
            ws.cell(row=idx, column=1, value=str(participant))
            ws.cell(row=idx, column=2, value=total_score)
            ws.cell(row=idx, column=3, value=total_ranks[participant])
            
            # Add dimension scores/ranks
            col_offset = 4
            for dim_num in correlations.keys():
                dim_data = dimension_results[dim_num]
                ws.cell(row=idx, column=col_offset, value=dim_data['scores'][participant])
                ws.cell(row=idx, column=col_offset + 1, value=dim_data['ranks'][participant])
                col_offset += 2
            
            for col in range(1, col_offset):
                ConstructValidityFormatter._apply_data_style(ws.cell(row=idx, column=col))
            
        # Adjust column widths
        ConstructValidityFormatter._adjust_column_widths(ws)
        
    @staticmethod
    def _apply_header_style(cell):
        """Apply consistent header styling"""
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        cell.border = Border(
            bottom=Side(style='medium'),
            top=Side(style='medium'),
            left=Side(style='thin'),
            right=Side(style='thin')
        )
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    @staticmethod
    def _apply_data_style(cell):
        """Apply consistent data cell styling"""
        cell.border = Border(
            bottom=Side(style='thin'),
            top=Side(style='thin'),
            left=Side(style='thin'),
            right=Side(style='thin')
        )
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    @staticmethod
    def _adjust_column_widths(ws):
        """Adjust column widths based on content"""
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = min(adjusted_width, 40)