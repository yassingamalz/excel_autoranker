# src/core/formatters/split_half_formatter.py
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from typing import Dict

class SplitHalfFormatter:
    """Formats Split-Half reliability results into Excel worksheet"""
    
    @staticmethod
    def format_results(wb: Workbook, results: Dict, prefix: str = "") -> None:
        """Format results to the active sheet with optional prefix"""
        ws = wb.active
        title_prefix = f"{prefix} " if prefix else ""
        ws.title = f"{title_prefix}Split Half - التجزئة النصفية"
        SplitHalfFormatter.format_results_to_sheet(ws, results)

    @staticmethod
    def format_results_to_sheet(ws: Worksheet, results: Dict) -> None:
        """Format results to a specific worksheet"""
        # Apply headers
        headers = [
            "Metric / المقياس",
            "Value / القيمة",
            "Interpretation / التفسير"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            SplitHalfFormatter._apply_header_style(cell)
        
        # Add results
        row_data = [
            ("Odd Questions Count / عدد الأسئلة الفردية", 
             len(results['odd_questions']), 
             ""),
            ("Even Questions Count / عدد الأسئلة الزوجية",
             len(results['even_questions']),
             ""),
            ("Pearson Correlation / معامل ارتباط بيرسون",
             round(results['pearson_correlation'], 6),
             ""),
            ("Spearman-Brown Coefficient / معامل سبيرمان-براون",
             round(results['spearman_brown'], 6),
             results.get('interpretation', ''))
        ]
        
        for row, (metric, value, interp) in enumerate(row_data, 2):
            ws.cell(row=row, column=1, value=metric)
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=3, value=interp)
            SplitHalfFormatter._apply_data_style(ws.cell(row=row, column=1))
            SplitHalfFormatter._apply_data_style(ws.cell(row=row, column=2))
            SplitHalfFormatter._apply_data_style(ws.cell(row=row, column=3))
            
        # Add participant sums table starting at row 8
        ws.cell(row=7, column=1, value="Participant Sums / مجموع درجات المشاركين")
        ws.cell(row=8, column=1, value="Participant / المشارك")
        ws.cell(row=8, column=2, value="Odd Sum / مجموع الأسئلة الفردية")
        ws.cell(row=8, column=3, value="Even Sum / مجموع الأسئلة الزوجية")
        
        for col in range(1, 4):
            SplitHalfFormatter._apply_header_style(ws.cell(row=8, column=col))
        
        odd_sums = results['odd_sums']
        even_sums = results['even_sums']
        
        for idx, (participant, odd_sum) in enumerate(odd_sums.items(), 9):
            ws.cell(row=idx, column=1, value=str(participant))
            ws.cell(row=idx, column=2, value=odd_sum)
            ws.cell(row=idx, column=3, value=even_sums[participant])
            
            for col in range(1, 4):
                SplitHalfFormatter._apply_data_style(ws.cell(row=idx, column=col))
            
        # Adjust column widths
        SplitHalfFormatter._adjust_column_widths(ws)
        
    @staticmethod
    def _apply_header_style(cell):
        """Apply consistent header styling"""
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        cell.border = Border(
            bottom=Side(style='medium'),
            top=Side(style='medium'),
            left=Side(style='thin'),
            right=Side(style='thin')
        )
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    @staticmethod
    def _apply_data_style(cell):
        """Apply consistent data cell styling"""
        cell.border = Border(
            bottom=Side(style='thin'),
            top=Side(style='thin'),
            left=Side(style='thin'),
            right=Side(style='thin')
        )
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    @staticmethod
    def _adjust_column_widths(ws):
        """Adjust column widths based on content"""
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = min(adjusted_width, 40)